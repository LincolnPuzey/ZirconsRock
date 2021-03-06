from openpyxl import *
from openpyxl.worksheet import *
from openpyxl.styles import *
from openpyxl.workbook import *
from openpyxl.utils import get_column_letter


def styleUPb(fileLocation):
    dirfile=str(fileLocation).replace('\\', '/').replace('\t', '/t').replace('\n', '/n').replace('\0', '/0')
    ogFileName = dirfile.split('/')[-1]

    main(fileLocation)


def addRows(wb1, ws1, shift, sheetName):
    ws2 = wb1.create_sheet('%s' % (sheetName))

    start_row = 1
    start_col = 1

    for row in ws1.iter_rows(min_row = start_row):
        for cell in row:
            ws2.cell(row = start_row + shift, column = start_col, value = cell.value)
            start_col += 1
        start_row += 1
        start_col = 1


def report(sheetName, wb, new_wb, bold, border):
    ws1 = wb['%s' % (sheetName)]
    addRows(new_wb, ws1, 2, sheetName)
    ws1_new = new_wb['%s' % (sheetName)]
    # print ws1_new.max_column
    ws1_new['A1'] = "U-Pb AGES (Ma)"
    ws1_new['A1'].font = bold
    ws1_new.merge_cells('G1:I1')
    ws1_new['G1'].font = bold
    for i in range (0, 9):
        ws1_new.cell(row = 3, column = i + 1).font = bold
        ws1_new.cell(row = 3, column = i + 1).border = border


def tbclc(sheetName, wb, new_wb, bold, border):
    ws1 = wb['%s' % (sheetName)]
    addRows(new_wb, ws1, 0, sheetName)
    ws1_new = new_wb['%s' % (sheetName)]

    ws1_new.merge_cells('A6:N6')
    ws1_new.merge_cells('B1:C1')
    ws1_new.merge_cells('J1:K1')
    ws1_new.merge_cells('M1:N1')

    ws1_new['A6'] = "These columns only to be used for data input (cut and paste)"
    ws1_new['A6'].alignment = Alignment(horizontal="center")

    alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    for i in range (0,14):
        ws1_new['%s6' % alphabet[i]].border = Border(top=Side(border_style="thin",color='000000'),
     bottom=Side(border_style= "thin" ,color='000000'))

    ws1_new['E4'] = ""
    ws1_new['C5'] = "1s"
    ws1_new['E4'] = ""
    ws1_new['E5'] = "1s"
    ws1_new['G4'] = ""
    ws1_new['G5'] = "1s"
    ws1_new['I4'] = ""
    ws1_new['I5'] = "1s"
    ws1_new['B3'] = "Measured Ratios"
    ws1_new['J3'] = "Raw Counts"
    ws1_new['M3'] = "Mean Raw CPS Data"
    ws1_new['D4'] = "Pb207/U235"
    ws1_new['F4'] = "Pb206/U238"

    ws1_new['A6'].fill = PatternFill("solid", fgColor="1BB912")

    columnE4 = []
    columnG4 = []

    i = 7
    while ws1_new['E'+str(i)].value != None:
        columnE4.append(ws1_new['E'+str(i)].value)
        i += 1

    i = 7
    while ws1_new['G'+str(i)].value != None:
        columnG4.append(ws1_new['G'+str(i)].value)
        i += 1

    for j in range(0,len(columnE4)):
        ws1_new['G' + str(j+7)].value = columnE4[j]

    for k in range(0,len(columnG4)):
        ws1_new['E' + str(k+7)].value = columnG4[k]


def column25(sheetName, wb, new_wb, bold):
    border = Border(top = Side(border_style = "double", color = '900000'))
    font = Font(name='Calibri', size=11, bold=True, color='00000000')
    font_heading = Font(name='Calibri', size=12, bold=True, color='900000')
    ws1 = wb['%s' % (sheetName)]
    addRows(new_wb, ws1, 1, sheetName)
    ws1_new = new_wb['%s' % (sheetName)]
    delete_column(ws1_new, 11)
    delete_column(ws1_new, 13)
    delete_column(ws1_new, 13)
    delete_column(ws1_new, 14)
    ws1_new.merge_cells('B2:I2')
    ws1_new.merge_cells('G1:I1')
    ws1_new.merge_cells('K2:L2')
    ws1_new.merge_cells('N2:U2')

    ws1_new['A2'] = ""
    ws1_new['N2'] = ""
    ws1_new['G1'] = "Common-Lead Corrected"
    ws1_new['G1'].font = Font(name='Calibri', size=12, bold=True, color='0000d4')
    ws1_new['B2'] = "RATIOS"
    ws1_new['B2'].font = font_heading
    ws1_new['B2'].alignment = Alignment(horizontal="center")
    ws1_new['K2'] = "CONCENTRATIONS (ppm)"
    ws1_new['K2'].font = font_heading
    ws1_new['K2'].alignment = Alignment(horizontal="center")
    ws1_new['O2'] = "AGES (Ma)"
    ws1_new['O2'].font = font_heading
    ws1_new['O2'].alignment = Alignment(horizontal="center")
    ws1_new['A3'].font = bold

    for i in range (0, 21):
        if i == 0 or i == 9 or i == 12:
            continue
        else:
            ws1_new.cell(row = 3, column = i + 1).font = font
            ws1_new.cell(row = 3, column = i + 1).border = border
    new_wb["%s" % sheetName].freeze_panes = "B1"

def column38(sheetName, wb, new_wb, bold):
    border = Border(top = Side(border_style = "double", color = '900000'))
    font = Font(name='Calibri', size=11, bold=True, color='00000000')
    font_heading = Font(name='Calibri', size=12, bold=True, color='900000')
    ws1 = wb['%s' % (sheetName)]
    addRows(new_wb, ws1, 1, sheetName)
    ws1_new = new_wb['%s' % (sheetName)]
    i=2
    zirconnumlist = []
    while ws1['N'+str(i)].value != None:
        zirconnumlist.append(ws1['N'+str(i)].value)
        i += 1

    for j in range(0,len(zirconnumlist)):
        ws1_new['AN' + str(j+3)].value = zirconnumlist[j]

    i=2
    sequenceNumList = []
    while ws1['O'+str(i)].value != None:
        sequenceNumList.append(ws1['O'+str(i)].value)
        i += 1

    for j in range(0,len(sequenceNumList)):
        ws1_new['AO' + str(j+3)].value = sequenceNumList[j]

    delete_column(ws1_new, 14)
    delete_column(ws1_new, 14)
    delete_column(ws1_new, 32)
    delete_column(ws1_new, 25)
    delete_column(ws1_new, 15)
    delete_column(ws1_new, 11)

    for i in range (0, 35):
        ws1_new.cell(row = 3, column = i + 1).font = bold

    ws1_new.merge_cells('B2:I2')
    ws1_new.merge_cells('G1:I1')
    ws1_new.merge_cells('K2:L2')
    ws1_new.merge_cells('N2:U2')
    ws1_new.merge_cells('W2:AA2')
    ws1_new.merge_cells('AC2:AF2')
    ws1_new.merge_cells('AH2:AI2')

    ws1_new['A2'] = ""
    ws1_new['G1'] = "Common-Lead Corrected"
    ws1_new['G1'].font = Font(name='Calibri', size=12, bold=True, color='0000d4')
    ws1_new['B2'] = "RATIOS"
    ws1_new['B2'].font = font_heading
    ws1_new['B2'].alignment = Alignment(horizontal="center")
    ws1_new['K2'] = "CONCENTRATIONS (ppm)"
    ws1_new['K2'].font = font_heading
    ws1_new['K2'].alignment = Alignment(horizontal="center")
    ws1_new['N2'] = "AGES (Ma)"
    ws1_new['N2'].font = font_heading
    ws1_new['N2'].alignment = Alignment(horizontal="center")
    ws1_new['W2'] = "NORMAL CONCORDIA PLOT DATA"
    ws1_new['W2'].font = font_heading
    ws1_new['W2'].alignment = Alignment(horizontal="center")
    ws1_new['AC2'] = "INVERSE CONCORDIA PLOT DATA"
    ws1_new['AC2'].font = font_heading
    ws1_new['AC2'].alignment = Alignment(horizontal="center")
    ws1_new['AH2'] = "SEQUENCE No."
    ws1_new['AH2'].font = font_heading
    ws1_new['AH2'].alignment = Alignment(horizontal="center")

    for i in range (0, 35):
        if i == 0 or i == 9 or i == 12 or i == 21 or i == 27 or i == 32:
            continue
        else:
            ws1_new.cell(row = 3, column = i + 1).font = font
            ws1_new.cell(row = 3, column = i + 1).border = border

    new_wb["%s" % sheetName].freeze_panes = "B1"

    if ws1_new['AC3'].value == "" and ws1_new['AD3'].value == "" and ws1_new['AE3'].value == "" and ws1_new['AF3'].value == "":
        delete_column(ws1_new, 29)
        delete_column(ws1_new, 29)
        delete_column(ws1_new, 29)
        delete_column(ws1_new, 29)
        delete_column(ws1_new, 29)
    
def delete_column(ws, delete_column):
    if isinstance(delete_column, str):
        delete_column = openpyxl.cell.column_index_from_string(delete_column)
    assert delete_column >= 1, "Column numbers must be 1 or greater"

    for column in range(delete_column, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).value = \
                    ws.cell(row=row, column=column+1).value


def main(fileLocation):
    wb = load_workbook(filename = '%s' % fileLocation)
    sheetNames = wb.get_sheet_names()

    new_wb = Workbook()
    bold = Font(b=True)
    border = Border(top = Side(border_style = "thin", color = '00000000'),
        bottom = Side(border_style = "double", color = "00000000"))

    for i in range (0, len(sheetNames)):
        if sheetNames[i] == "Ratios raw" or sheetNames[i] == "Ages raw":
            sheetName = sheetNames[i]
            ws1 = wb['%s' % (sheetName)]
            addRows(new_wb, ws1, 0, sheetName)

        elif sheetNames[i] == "Report":
            report(sheetNames[i], wb, new_wb,bold, border)

        elif sheetNames[i] == "ToBeCommonLeadCorrected":
            tbclc(sheetNames[i], wb, new_wb,bold, border)
        else:
            sheetName = sheetNames[i]
            ws1 = wb['%s' % (sheetName)]
            maxColumns = ws1.max_column

            if maxColumns == 25:
                column25(sheetName, wb, new_wb, bold)

            else:
                column38(sheetName, wb, new_wb, bold)
    # completed = False

    # while ~completed:
    #   try:
    for sheet in new_wb.worksheets:
        if sheet.title == 'Sheet':
            new_wb.remove_sheet(sheet)
            
    new_wb.save('%s' % fileLocation)
    # os.remove('%s' % fileLocation)
    #       completed = True
    #   except:
    #       print "Please close the final.xlsx file before continuing..."

    #Deletes the original xlsx file
